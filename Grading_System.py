import os
import tkinter as tk
import customtkinter as ctk
import openpyxl
from tkinter import ttk, messagebox

# refresh the data Tables & Labels
def refresh_gui():
    refresh_interval = 100
    treeview.after(refresh_interval, display_excel_data)

# shortcut key for refresh F5
def on_key_press(event):
    # Check if the F5 key is pressed
    if event.keysym == "F5":
        # Call the refresh_gui function
        refresh_gui()

filename = "filename.xlsx"
headings = ["No.", "Student Name", "Quiz1", "Perf1", "Midterm Project", "Midterm Exam", "Quiz2", "Perf2",
            "Finalterm Project", "Finalterm Exam", "Initial Grade", "Final Grade", "Total Grade", "Status",
            "Course", "Year", "Section", "Remarks", "Teacher Name", "Subject", "Semester", "Term", "MTerm-Q1", 
            "MTerm-Q2","MTerm-Q3", "MTerm-P1", "MTerm-P2", "MTerm-P3", "MTerm-A1", "FTerm-Q1", "FTerm-Q2",
            "FTerm-Q3", "FTerm-P1", "FTerm-P2", "FTerm-P3", "FTerm-A2"]

selected_row_index = -1

def calculate_score(values):
    MTerm = []
    for value in values:
        MTerm.append(int(value or 0))
    score = sum(MTerm)
    return score

def calculate_initial_grade(row):
    quiz = ((int(row[2] or 0) / 30) * 100) * 0.25
    perf = ((int(row[3] or 0) / 150) * 100) * 0.25
    mid_project = ((int(row[4] or 0) / 100) * 100) * 0.25
    mid_exam = ((int(row[5] or 0) / 50) * 100) * 0.25
    initial_grade = quiz + perf + mid_project + mid_exam
    return initial_grade

def calculate_final_grade(row):
    f_quiz = ((int(row[6] or 0) / 30) * 100) * 0.25
    f_perf = ((int(row[7] or 0) / 150) * 100) * 0.25
    final_project = ((int(row[8] or 0) / 100) * 100) * 0.25
    final_exam = ((int(row[9] or 0) / 50) * 100) * 0.25
    final_grade = f_quiz + f_perf + final_project + final_exam
    return final_grade

def calculate_total_grade(row):
    midterm = int(row[10] or 0)
    finalterm = int(row[11] or 0)
    grade = [midterm,finalterm]
    total_grade = sum(grade) / 2
    return total_grade

def display_excel_data():
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    treeview.delete(*treeview.get_children())
    rows = sheet.iter_rows(values_only=True)
    next(rows)
    for i, row in enumerate(rows, start=1):
        quiz_score = calculate_score([row[22], row[23], row[24]])
        perf_score = calculate_score([row[25], row[26], row[27]])
        quiz_score2 = calculate_score([row[29], row[30], row[31]])
        perf_score2 = calculate_score([row[32], row[33], row[34]])
        initial = calculate_initial_grade(row)
        final_grade_total = calculate_final_grade(row)
        total_grade = calculate_total_grade(row)
        remarks = "Passed" if total_grade >= 75 else "Failed"
        values = (i,) + row[1:2] + (quiz_score,) + row[3:3] + (perf_score,) + row[4:] + (quiz_score2,) + row[7:7] + (perf_score2,) + row[8:] + (initial,) + row[11:11] + (final_grade_total,) + row[12:12] + (total_grade,) + row[13:]+ (remarks,) + row[18:]
        treeview.insert("", "end", values=values)
        insert_scores_into_excel(sheet, i + 1, quiz_score, perf_score, quiz_score2, perf_score2, initial, final_grade_total, total_grade,remarks)
    workbook.save(filename)
    workbook.close()

def insert_scores_into_excel(sheet, row_index, quiz_score, perf_score, quiz_score2, perf_score2, initial, final_grade_total, total_grade,remarks):
    sheet.cell(row=row_index, column=3).value = quiz_score
    sheet.cell(row=row_index, column=4).value = perf_score
    sheet.cell(row=row_index, column=7).value = quiz_score2
    sheet.cell(row=row_index, column=8).value = perf_score2
    sheet.cell(row=row_index, column=11).value = initial
    sheet.cell(row=row_index, column=12).value = final_grade_total
    sheet.cell(row=row_index, column=13).value = total_grade
    sheet.cell(row=row_index, column=18).value = remarks


def create_excel_file():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(headings)
    workbook.save(filename)
    print(f"Excel file '{filename}' created successfully.")

def get_last_no():
    if not os.path.exists(filename):
        return 0
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    last_row = sheet.max_row
    if last_row == 1:
        return 0
    else:
        return sheet.cell(row=last_row, column=1).value
    
def cancel_operation(cancel_id):
    if cancel_id == 1:
        Prof_edit.grid_remove()
        Professor_details.grid(row=0, column=0, padx=15,pady=(5,0), sticky="nsew")
    elif cancel_id == 2:
        details_edit.grid_remove()
        Student_details.grid(row=2, column=0, padx=10, pady=10, sticky="nsew")
    elif cancel_id == 3:
        edit_Activity.grid_remove()
        Student_Activity.grid(row=3, column=0, padx=10, pady=10, sticky="nsew")
    elif cancel_id == 4:
        edit_Activity2.grid_remove()
        Student_Activity2.grid(row=3, column=0, padx=10, pady=10, sticky="nsew")
    
    if selected_row_index >= 0:
        selected_item = treeview.get_children()[selected_row_index]
        treeview.selection_set(selected_item)
        treeview.focus(selected_item)

def add_data(name,status,course,year,section):
    last_no = get_last_no()
    if last_no is None:
        last_no = 0
    new_no = last_no + 1
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    data = [new_no,name,"","","","","","","","","","","",status,course,year,section]
    sheet.append(data)
    workbook.save(filename)
    print("Data added successfully.")
    display_excel_data()

def delete_data(row_number):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    sheet.delete_rows(row_number + 2)

    for row in range(row_number + 2, sheet.max_row + 1):
        sheet.cell(row=row, column=1).value = row - 1
    workbook.save(filename)
    print(f"Row {row_number+1} deleted successfully.")
    display_excel_data()

def edit_data(row_number, column_number, new_value):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    sheet.cell(row=row_number + 2, column=column_number + 1).value = new_value
    workbook.save(filename)
    print(f"Data in row {row_number + 1}, column {column_number} edited successfully.")
    display_excel_data()

def on_select(event, treeview):
    global selected_row_index
    selection = treeview.selection()
    if selection:
        
        selected_item = treeview.item(selection[0])
        values = selected_item['values']
        selected_row_index = int(treeview.index(selection[0]))
        
        Student_name_label.configure(text=values[1])
        FirstQuarter_label.configure(text=values[10])
        SecQuarter_label.configure(text=values[11])
        Final_grade_label.configure(text=values[12])
        Student_Status_label.configure(text=values[13])
        Course_label.configure(text=values[14])
        Year_label.configure(text=values[15])
        Section_label.configure(text=values[16])
        Remarks_label.configure(text=values[17])
        
        teacher_name_label.configure(text=values[18])
        subject_name_label.configure(text=values[19])
        semester_label.configure(text=str(values[20]) +" | "+ str(values[21]))
        
        q1.configure(text=values[22])
        q2.configure(text=values[23])
        q3.configure(text=values[24])
        p1.configure(text=values[25])
        p2.configure(text=values[26])
        p3.configure(text=values[27])
        mp.configure(text=values[4])
        me.configure(text=values[5])
        A1.configure(text=values[28])
        
        q1f.configure(text=values[29])
        q2f.configure(text=values[30])
        q3f.configure(text=values[31])
        p1f.configure(text=values[32])
        p2f.configure(text=values[33])
        p3f.configure(text=values[34])
        fp.configure(text=values[8])
        fe.configure(text=values[9])
        A1f.configure(text=values[35])
        
        treeview.tag_configure('highlight', background='yellow')

def student_delete():
    global selected_row_index
    if selected_row_index >= 0:
        confirmation = messagebox.askyesno("Delete Confirmation", "Are you sure you want to delete this record?")
        if confirmation:
            delete_data(selected_row_index)
            selected_row_index = -1
            Student_name_label.configure(text="N/A")
        else:
            messagebox.showinfo("Delete Cancelled", "Record deletion has been cancelled.")
    else:
        messagebox.showinfo("No Selection", "No record selected for deletion.")
def professor_edit():
    global teacher_entry,subject_entry,semester_entry,term_entry,Prof_edit
    selected_item_id = treeview.focus()
    if selected_item_id:
        selected_row_index = int(treeview.index(selected_item_id))
        if selected_row_index >= 0:
            Prof_edit = ctk.CTkFrame(left_side)
            Prof_edit.grid(row=0, column=0, padx=10,pady=10, sticky="nsew")

            teacher_details = ctk.CTkLabel(Prof_edit, text="TEACHER DETAILS")
            teacher_details.grid(row=0, column=0, sticky="w",padx=10)
            save_prof = ctk.CTkButton(Prof_edit, text="SAVE", command=lambda: save_data(1))
            save_prof.grid(row=5, column=1, sticky="nsew", pady=(5,10))
            cancel_prof = ctk.CTkButton(Prof_edit, text="CANCEL", command=lambda: cancel_operation(1))
            cancel_prof.grid(row=0, column=1, sticky="nsew",pady=(10,5))

            teacher_name = ctk.CTkLabel(Prof_edit, text="TEACHER NAME: ")
            teacher_subj = ctk.CTkLabel(Prof_edit, text="SUBJECT: ")
            teacher_sem = ctk.CTkLabel(Prof_edit, text="SEMESTER: ")
            teacher_term = ctk.CTkLabel(Prof_edit, text="TERM: ")

            teacher_name.grid(row=1, column=0, sticky="w", padx=(10,0))
            teacher_subj.grid(row=2, column=0, sticky="w", padx=(10,0))
            teacher_sem.grid(row=3, column=0, sticky="w", padx=(10,0))
            teacher_term.grid(row=4, column=0, sticky="w", padx=(10,0))

            teacher_entry = ctk.CTkEntry(Prof_edit)
            subject_entry = ctk.CTkEntry(Prof_edit)
            semester_entry = ctk.CTkEntry(Prof_edit)
            term_entry = ctk.CTkEntry(Prof_edit)
            
            teacher_entry.grid(row=1, column=1, sticky="w")
            subject_entry.grid(row=2, column=1, sticky="w")
            semester_entry.grid(row=3, column=1, sticky="w")
            term_entry.grid(row=4, column=1, sticky="w")
            
            selected_item = treeview.item(selected_item_id)
            values = selected_item['values']
            teacher_entry.insert(0, values[18])
            subject_entry.insert(0, values[19])
            semester_entry.insert(0, values[20])
            term_entry.insert(0, values[21])
    else:
        messagebox.showinfo("No Selection", "No record selected for editing info.")
          
def student_edit():
    global Name_entry,Status_option,Course_entry,Year_entry,Section_entry,details_edit,Status_entry
    selected_item_id = treeview.focus()
    if selected_item_id:
        selected_row_index = int(treeview.index(selected_item_id))
        if selected_row_index >= 0:
            Student_details.grid_remove()
            details_edit = ctk.CTkFrame(left_side)
            details_edit.grid(row=2, column=0, padx=10, pady=10, sticky="nsew")
            
            Status_entry = ctk.CTkComboBox(details_edit)
            Status_entry.configure(values = ['Regular Student', 'Irregular Student'])
            Status_entry.set('Regular Student')
            
            Student_name = ctk.CTkLabel(details_edit, text="STUDENT NAME: ")
            Status_name = ctk.CTkLabel(details_edit, text="STUDENT STATUS: ")
            Course = ctk.CTkLabel(details_edit, text="COURSE: ")
            Year = ctk.CTkLabel(details_edit, text="YEAR: ")
            Section = ctk.CTkLabel(details_edit, text="SECTION: ")
            
            Student_name.grid(row=1, column=0, sticky="w",padx=10)
            Status_name.grid(row=2, column=0, sticky="w",padx=10)
            Course.grid(row=3, column=0, sticky="w",padx=10)
            Year.grid(row=4, column=0, sticky="w",padx=10)
            Section.grid(row=5, column=0, sticky="w",padx=10)

            add_button = ctk.CTkButton(details_edit, text="SAVE", command=lambda: save_data(2))
            cancel_button = ctk.CTkButton(details_edit, text="CANCEL", command=lambda: cancel_operation(2))
            add_button.grid(row=6, column=0, sticky="w",padx=(10,10),pady=(0,10))
            cancel_button.grid(row=6, column=1, sticky="w",pady=(0,10))
            
            Name_entry = ctk.CTkEntry(details_edit)
            Course_entry = ctk.CTkEntry(details_edit)
            Year_entry = ctk.CTkEntry(details_edit)
            Section_entry = ctk.CTkEntry(details_edit)
            Name_entry.grid(row=1, column=1, sticky="nsew",pady=(10,0))
            Status_entry.grid(row=2, column=1, sticky="nsew")
            Course_entry.grid(row=3, column=1, sticky="nsew")
            Year_entry.grid(row=4, column=1, sticky="nsew")
            Section_entry.grid(row=5, column=1, sticky="nsew",pady=(0,10))
            
            selected_item = treeview.item(selected_item_id)
            values = selected_item['values']
            Name_entry.insert(0, values[1])
            Course_entry.insert(0, values[14])
            Year_entry.insert(0, values[15])
            Section_entry.insert(0, values[16])
    else:
        messagebox.showinfo("No Selection", "No record selected for editing info.")
        
def student_add():
    global Name_entry,Status_option,Course_entry,Year_entry,Section_entry,details_edit
    Student_details.grid_remove()
    details_edit = ctk.CTkFrame(left_side)
    details_edit.grid(row=2, column=0, padx=10, pady=10, sticky="nsew")
    
    Status_option = tk.StringVar()
    Status_entry = ttk.Combobox(details_edit, textvariable=Status_option)
    Status_entry['values'] = ('Regular Student', 'Irregular Student')
    Status_entry.current(0)
    
    Student_name = ctk.CTkLabel(details_edit, text="STUDENT NAME: ")
    Status_name = ctk.CTkLabel(details_edit, text="STUDENT STATUS: ")
    Course = ctk.CTkLabel(details_edit, text="COURSE: ")
    Year = ctk.CTkLabel(details_edit, text="YEAR: ")
    Section = ctk.CTkLabel(details_edit, text="SECTION: ")
    
    Student_name.grid(row=1, column=0, sticky="w",padx=10)
    Status_name.grid(row=2, column=0, sticky="w",padx=10)
    Course.grid(row=3, column=0, sticky="w",padx=10)
    Year.grid(row=4, column=0, sticky="w",padx=10)
    Section.grid(row=5, column=0, sticky="w",padx=10)

    add_button = ctk.CTkButton(details_edit, text="SAVE", command=save_add)
    cancel_button = ctk.CTkButton(details_edit, text="CANCEL", command=lambda: cancel_operation(2))
    add_button.grid(row=6, column=0, sticky="w",padx=(10,0),pady=(0,10))
    cancel_button.grid(row=6, column=1, sticky="w",pady=(0,10),padx=(10,0))
    
    Name_entry = ctk.CTkEntry(details_edit)
    Course_entry = ctk.CTkEntry(details_edit)
    Year_entry = ctk.CTkEntry(details_edit)
    Section_entry = ctk.CTkEntry(details_edit)
    Name_entry.grid(row=1, column=1, sticky="nsew",pady=(10,0),padx=(10,0))
    Status_entry.grid(row=2, column=1, sticky="nsew",padx=(10,0))
    Course_entry.grid(row=3, column=1, sticky="nsew",padx=(10,0))
    Year_entry.grid(row=4, column=1, sticky="nsew",padx=(10,0))
    Section_entry.grid(row=5, column=1, sticky="nsew",padx=(10,0))

def save_add():
    new_name = Name_entry.get()
    new_status = Status_option.get()
    new_course = Course_entry.get()
    new_year = Year_entry.get()
    new_section = Section_entry.get()
    if not new_name:
        messagebox.showwarning("Empty Name", "Name cannot be empty. Please enter a name.")
        return
    elif not new_status:
        messagebox.showwarning("Empty Status", "Course cannot be empty. Please enter a Status.")
        return
    elif not new_course:
        messagebox.showwarning("Empty Course", "Course cannot be empty. Please enter a Course.")
        return
    elif not new_year:
        messagebox.showwarning("Empty Year", "Year cannot be empty. Please enter a Year.")
        return
    elif not new_section:
        messagebox.showwarning("Empty Section", "Name cannot be empty. Please enter a Section.")
        return
    add_data(new_name,new_status,new_course,new_year,new_section)
    details_edit.grid_remove()
    Student_details.grid(row=2, column=0, padx=10, pady=10, sticky="nsew")
    
def activity_edit_midterm():
    global edit_Activity,q1_entry,q2_entry,q3_entry,p1_entry,p2_entry,p3_entry,A1_entry,mp_entry,me_entry
    selected_item_id = treeview.focus()
    if selected_item_id:
        selected_row_index = int(treeview.index(selected_item_id))
        if selected_row_index >= 0:
            Student_Activity.grid_remove()
            edit_Activity = ctk.CTkFrame(left_side)
            edit_Activity.grid(row=3, column=0, padx=10, pady=10, sticky="nsew")

            details_title3 = ctk.CTkFrame(edit_Activity)
            details_title3.grid(row=0, column=0, sticky="w", padx=10)
            student_act = ctk.CTkLabel(details_title3, text="STUDENT ACTIVITIES | MIDTERM")
            student_act.grid(row=0, column=0, sticky="w")
            student_term = ctk.CTkButton(edit_Activity, text="CANCEL", command=lambda: cancel_operation(3))
            student_term.grid(row=0, column=1, sticky="w", padx=(0,10), pady=(0,10))
            edit_act = ctk.CTkButton(edit_Activity, text="SAVE ACTIVITIES", command=lambda: save_data(3))
            edit_act.grid(row=0, column=2, sticky="nsew", pady=(0,10))
            Quiz1 = ctk.CTkLabel(edit_Activity, text="QUIZ #1: ")
            Quiz2 = ctk.CTkLabel(edit_Activity, text="QUIZ #2: ")
            Quiz3 = ctk.CTkLabel(edit_Activity, text="QUIZ #3: ")
            Performance1 = ctk.CTkLabel(edit_Activity, text="PERFORMANCE #1: ")
            Performance2 = ctk.CTkLabel(edit_Activity, text="PERFORMANCE #2: ")
            Performance3 = ctk.CTkLabel(edit_Activity, text="PERFORMANCE #3: ")
            MIDTERM_P = ctk.CTkLabel(edit_Activity, text="MIDTERM PROJECT: ")
            MIDTERM_E = ctk.CTkLabel(edit_Activity, text="MIDTERM EXAM: ")
            Attendance_Midterm = ctk.CTkLabel(edit_Activity, text="ATTENDANCE MIDTERM: ")

            Quiz1.grid(row=1, column=0, sticky="w", padx=10)
            Quiz2.grid(row=2, column=0, sticky="w", padx=10)
            Quiz3.grid(row=3, column=0, sticky="w", padx=10)
            Performance1.grid(row=4, column=0, sticky="w", padx=10)
            Performance2.grid(row=5, column=0, sticky="w", padx=10)
            Performance3.grid(row=6, column=0, sticky="w", padx=10)
            MIDTERM_P.grid(row=7, column=0, sticky="w", padx=10)
            MIDTERM_E.grid(row=8, column=0, sticky="w", padx=10)
            Attendance_Midterm.grid(row=9, column=0, sticky="w", padx=10)

            # Activities Entry
            q1_entry = ctk.CTkEntry(edit_Activity)
            q2_entry = ctk.CTkEntry(edit_Activity)
            q3_entry = ctk.CTkEntry(edit_Activity)
            p1_entry = ctk.CTkEntry(edit_Activity)
            p2_entry = ctk.CTkEntry(edit_Activity)
            p3_entry = ctk.CTkEntry(edit_Activity)
            mp_entry = ctk.CTkEntry(edit_Activity)
            me_entry = ctk.CTkEntry(edit_Activity)
            A1_entry = ctk.CTkEntry(edit_Activity)

            q1_entry.grid(row=1, column=1, sticky="w")
            q2_entry.grid(row=2, column=1, sticky="w")
            q3_entry.grid(row=3, column=1, sticky="w")
            p1_entry.grid(row=4, column=1, sticky="w")
            p2_entry.grid(row=5, column=1, sticky="w")
            p3_entry.grid(row=6, column=1, sticky="w")
            mp_entry.grid(row=7, column=1, sticky="w")
            me_entry.grid(row=8, column=1, sticky="w")
            A1_entry.grid(row=9, column=1, sticky="w")
            
            tq1_entry = ctk.CTkLabel(edit_Activity, text=" / 10")
            tq2_entry = ctk.CTkLabel(edit_Activity, text=" / 10")
            tq3_entry = ctk.CTkLabel(edit_Activity, text=" / 10")
            tp1_entry = ctk.CTkLabel(edit_Activity, text=" / 50")
            tp2_entry = ctk.CTkLabel(edit_Activity, text=" / 50")
            tp3_entry = ctk.CTkLabel(edit_Activity, text=" / 50")
            tmp_entry = ctk.CTkLabel(edit_Activity, text=" / 100")
            tme_entry = ctk.CTkLabel(edit_Activity, text=" / 50")
            tA1_entry = ctk.CTkLabel(edit_Activity, text=" / 30")

            tq1_entry.grid(row=1, column=2, sticky="w")
            tq2_entry.grid(row=2, column=2, sticky="w")
            tq3_entry.grid(row=3, column=2, sticky="w")
            tp1_entry.grid(row=4, column=2, sticky="w")
            tp2_entry.grid(row=5, column=2, sticky="w")
            tp3_entry.grid(row=6, column=2, sticky="w")
            tmp_entry.grid(row=7, column=2, sticky="w")
            tme_entry.grid(row=8, column=2, sticky="w")
            tA1_entry.grid(row=9, column=2, sticky="w")
            
            selected_item = treeview.item(selected_item_id)
            values = selected_item['values']
            q1_entry.insert(0, values[22])
            q2_entry.insert(0, values[23])
            q3_entry.insert(0, values[24])
            p1_entry.insert(0, values[25])
            p2_entry.insert(0, values[26])
            p3_entry.insert(0, values[27])
            mp_entry.insert(0, values[4])
            me_entry.insert(0, values[5])
            A1_entry.insert(0, values[28])
            
            
    else:
        messagebox.showinfo("No Selection", "No record selected for editing activity.")
        
def activity_edit_final():
    global edit_Activity2,q1f_entry,q2f_entry,q3f_entry,p1f_entry,p2f_entry,p3f_entry,fp_entry,fe_entry,A1f_entry,tq1f_entry,tq2f_entry,tq3f_entry,tp1f_entry,tp2f_entry,tp3f_entry,tfp_entry,tfe_entry,tA1f_entry
    selected_item_id = treeview.focus()
    if selected_item_id:
        selected_row_index = int(treeview.index(selected_item_id))
        if selected_row_index >= 0:
            Student_Activity.grid_remove()
            edit_Activity2 = ctk.CTkFrame(left_side)
            edit_Activity2.grid(row=3, column=0, padx=10, pady=10, sticky="nsew")

            details_title4 = ctk.CTkFrame(edit_Activity2)
            details_title4.grid(row=0, column=0, sticky="w", padx=10)
            student_act = ctk.CTkLabel(details_title4, text="STUDENT ACTIVITIES | FINAL-TERM")
            student_act.grid(row=0, column=0, sticky="w")
            student_term = ctk.CTkButton(edit_Activity2, text="CANCEL", command=lambda: cancel_operation(4))
            student_term.grid(row=0, column=1, sticky="w", padx=(0,10), pady=(0,10))
            edit_act = ctk.CTkButton(edit_Activity2, text="SAVE ACTIVITIES", command=lambda: save_data(4))
            edit_act.grid(row=0, column=2, sticky="nsew", pady=(0,10))
            Quiz1 = ctk.CTkLabel(edit_Activity2, text="QUIZ #1: ")
            Quiz2 = ctk.CTkLabel(edit_Activity2, text="QUIZ #2: ")
            Quiz3 = ctk.CTkLabel(edit_Activity2, text="QUIZ #3: ")
            Performance1 = ctk.CTkLabel(edit_Activity2, text="PERFORMANCE #1: ")
            Performance2 = ctk.CTkLabel(edit_Activity2, text="PERFORMANCE #2: ")
            Performance3 = ctk.CTkLabel(edit_Activity2, text="PERFORMANCE #3: ")
            MIDTERM_P = ctk.CTkLabel(edit_Activity2, text="MIDTERM PROJECT: ")
            MIDTERM_E = ctk.CTkLabel(edit_Activity2, text="MIDTERM EXAM: ")
            Attendance_Midterm = ctk.CTkLabel(edit_Activity2, text="ATTENDANCE MIDTERM: ")

            Quiz1.grid(row=1, column=0, sticky="w", padx=10)
            Quiz2.grid(row=2, column=0, sticky="w", padx=10)
            Quiz3.grid(row=3, column=0, sticky="w", padx=10)
            Performance1.grid(row=4, column=0, sticky="w", padx=10)
            Performance2.grid(row=5, column=0, sticky="w", padx=10)
            Performance3.grid(row=6, column=0, sticky="w", padx=10)
            MIDTERM_P.grid(row=7, column=0, sticky="w", padx=10)
            MIDTERM_E.grid(row=8, column=0, sticky="w", padx=10)
            Attendance_Midterm.grid(row=9, column=0, sticky="w", padx=10)

            # Activities Entry
            q1f_entry = ctk.CTkEntry(edit_Activity2)
            q2f_entry = ctk.CTkEntry(edit_Activity2)
            q3f_entry = ctk.CTkEntry(edit_Activity2)
            p1f_entry = ctk.CTkEntry(edit_Activity2)
            p2f_entry = ctk.CTkEntry(edit_Activity2)
            p3f_entry = ctk.CTkEntry(edit_Activity2)
            fp_entry = ctk.CTkEntry(edit_Activity2)
            fe_entry = ctk.CTkEntry(edit_Activity2)
            A1f_entry = ctk.CTkEntry(edit_Activity2)

            q1f_entry.grid(row=1, column=1, sticky="w")
            q2f_entry.grid(row=2, column=1, sticky="w")
            q3f_entry.grid(row=3, column=1, sticky="w")
            p1f_entry.grid(row=4, column=1, sticky="w")
            p2f_entry.grid(row=5, column=1, sticky="w")
            p3f_entry.grid(row=6, column=1, sticky="w")
            fp_entry.grid(row=7, column=1, sticky="w")
            fe_entry.grid(row=8, column=1, sticky="w")
            A1f_entry.grid(row=9, column=1, sticky="w")
            
            tq1f_entry = ctk.CTkLabel(edit_Activity2, text=" / 10")
            tq2f_entry = ctk.CTkLabel(edit_Activity2, text=" / 10")
            tq3f_entry = ctk.CTkLabel(edit_Activity2, text=" / 10")
            tp1f_entry = ctk.CTkLabel(edit_Activity2, text=" / 50")
            tp2f_entry = ctk.CTkLabel(edit_Activity2, text=" / 50")
            tp3f_entry = ctk.CTkLabel(edit_Activity2, text=" / 50")
            tfp_entry = ctk.CTkLabel(edit_Activity2, text=" / 100")
            tfe_entry = ctk.CTkLabel(edit_Activity2, text=" / 50")
            tA1f_entry = ctk.CTkLabel(edit_Activity2, text=" / 30")

            tq1f_entry.grid(row=1, column=2, sticky="w")
            tq2f_entry.grid(row=2, column=2, sticky="w")
            tq3f_entry.grid(row=3, column=2, sticky="w")
            tp1f_entry.grid(row=4, column=2, sticky="w")
            tp2f_entry.grid(row=5, column=2, sticky="w")
            tp3f_entry.grid(row=6, column=2, sticky="w")
            tfp_entry.grid(row=7, column=2, sticky="w")
            tfe_entry.grid(row=8, column=2, sticky="w")
            tA1f_entry.grid(row=9, column=2, sticky="w")
            
            selected_item = treeview.item(selected_item_id)
            values = selected_item['values']
            q1f_entry.insert(0, values[29])
            q2f_entry.insert(0, values[30])
            q3f_entry.insert(0, values[31])
            p1f_entry.insert(0, values[32])
            p2f_entry.insert(0, values[33])
            p3f_entry.insert(0, values[34])
            fp_entry.insert(0, values[8])
            fe_entry.insert(0, values[9])
            A1f_entry.insert(0, values[35])
    else:
        messagebox.showinfo("No Selection", "No record selected for editing activity.")

def save_data(save_id):
    if save_id == 1:
        if selected_row_index >= 0:
            new_data1 = teacher_entry.get()
            new_data2 = subject_entry.get()
            new_data3 = semester_entry.get()
            new_data4 = term_entry.get()
            edit_data(selected_row_index, 18, new_data1)
            edit_data(selected_row_index, 19, new_data2)
            edit_data(selected_row_index, 20, new_data3)
            edit_data(selected_row_index, 21, new_data4)
        Prof_edit.grid_remove()
        Professor_details.grid(row=0, column=0, padx=15,pady=(5,0), sticky="nsew")
    elif save_id == 2:
        if selected_row_index >= 0:
            new_data1 = Name_entry.get()
            new_data2 = Status_entry.get()
            new_data3 = Course_entry.get()
            new_data4 = Year_entry.get()
            new_data5 = Section_entry.get()
            edit_data(selected_row_index, 1, new_data1)
            edit_data(selected_row_index, 13, new_data2)
            edit_data(selected_row_index, 14, new_data3)
            edit_data(selected_row_index, 15, new_data4)
            edit_data(selected_row_index, 16, new_data5)
        details_edit.grid_remove()
        Student_details.grid(row=2, column=0, padx=10, pady=10, sticky="nsew")
    elif save_id == 3:
        if selected_row_index >= 0:
            new_data1 = q1_entry.get()
            new_data2 = q2_entry.get()
            new_data3 = q3_entry.get()
            new_data4 = p1_entry.get()
            new_data5 = p2_entry.get()
            new_data6 = p3_entry.get()
            new_data7 = mp_entry.get()
            new_data8 = me_entry.get()
            new_data9 = A1_entry.get()
            edit_data(selected_row_index, 22, new_data1)
            edit_data(selected_row_index, 23, new_data2)
            edit_data(selected_row_index, 24, new_data3)
            edit_data(selected_row_index, 25, new_data4)
            edit_data(selected_row_index, 26, new_data5)
            edit_data(selected_row_index, 27, new_data6)
            edit_data(selected_row_index, 4, new_data7)
            edit_data(selected_row_index, 5, new_data8)
            edit_data(selected_row_index, 28, new_data9)
        edit_Activity.grid_remove()
        Student_Activity.grid(row=3, column=0, padx=10, pady=10, sticky="nsew")
    elif save_id == 4:
        if selected_row_index >= 0:
            new_data1 = q1f_entry.get()
            new_data2 = q2f_entry.get()
            new_data3 = q3f_entry.get()
            new_data4 = p1f_entry.get()
            new_data5 = p2f_entry.get()
            new_data6 = p3f_entry.get()
            new_data7 = fp_entry.get()
            new_data8 = fe_entry.get()
            new_data9 = A1f_entry.get()
            edit_data(selected_row_index, 29, new_data1)
            edit_data(selected_row_index, 30, new_data2)
            edit_data(selected_row_index, 31, new_data3)
            edit_data(selected_row_index, 32, new_data4)
            edit_data(selected_row_index, 33, new_data5)
            edit_data(selected_row_index, 34, new_data6)
            edit_data(selected_row_index, 8, new_data7)
            edit_data(selected_row_index, 9, new_data8)
            edit_data(selected_row_index, 35, new_data9)
        edit_Activity2.grid_remove()
        Student_Activity.grid(row=3, column=0, padx=10, pady=10, sticky="nsew")
        
def alert_notif():
    messagebox.showinfo("Does not function Yet", "SAAAANDAALIIIII HINDI PAKO TAPOS DITO.\n OR TO BE CONTINUE NALANG!.. HEHE!")
    
def show_formula():
    messagebox.showinfo("Help", "The Formula that I use here was\nSum all Quizzes then Divide in Total of Quizzes\nSame goes to Performance, Mid and Final Term  Project, Mid and Final Term Exam\nThen multiply by 100 and multiply again by 0.25\nthe Q, P, MP, ME and so on and the result of every activity should be sum all of them.")

def toggle_theme():
    if check_white.get():
        print("Setting theme to light")
        check_white.set(True)
        file_menu.configure(bg="#DBDBDB", fg="#000000")
        theme_menu.configure(bg="#DBDBDB", fg="#000000")
        ctk.set_appearance_mode("light")
    else:
        print("Setting theme to dark")
        check_white.set(False)
        file_menu.configure(bg="#DBDBDB", fg="#000000")
        theme_menu.configure(bg="#DBDBDB", fg="#000000")
        ctk.set_appearance_mode("dark")

if os.path.exists(filename):
    print(f"Excel file '{filename}' already exists.")
else:
    create_excel_file()

root = tk.Tk()
root.title("Grading System")
root.geometry("1485x825")
root.bind("<KeyPress>", on_key_press)

menu_bar = tk.Menu(root)
root.configure(menu=menu_bar)

file_menu = tk.Menu(menu_bar, tearoff=0, bg="#DBDBDB", fg="#000000")
file_menu.add_command(label="Refresh       \"F5\"", command=refresh_gui)
file_menu.add_command(label="Formula", command=show_formula)
file_menu.add_command(label="Open",command=alert_notif)
file_menu.add_command(label="Save",command=alert_notif)
file_menu.add_separator()
file_menu.add_command(label="Exit", command=root.quit)

theme_menu = tk.Menu(menu_bar, tearoff=0, bg="#DBDBDB", fg="#000000")
check_white = tk.BooleanVar()
check_white.set(False)
menu_bar.add_cascade(label="File", menu=file_menu)
menu_bar.add_cascade(label="Theme", menu=theme_menu)

theme_menu.add_checkbutton(label="White/Black Theme", variable=check_white,command=toggle_theme)

frame = ctk.CTkFrame(root)
frame.pack(fill="both", expand=True)

left_side = ctk.CTkFrame(frame)
left_side.grid(row=0, column=0, sticky="nsew")

# Professor Details Label
Professor_details = ctk.CTkFrame(left_side)
Professor_details.grid(row=0, column=0, padx=10,pady=10, sticky="nsew")
details_title1 = ctk.CTkFrame(Professor_details)
details_title1.grid(row=0, column=0, sticky="w", padx=10,pady=10)
teacher_details = ctk.CTkLabel(details_title1, text="TEACHER DETAILS")
teacher_details.grid(row=0, column=0, sticky="w", padx=10)
edit_prof = ctk.CTkButton(Professor_details, text="EDIT", command=professor_edit)
edit_prof.grid(row=0, column=1, sticky="w")

teacher_name = ctk.CTkLabel(Professor_details, text="TEACHER NAME: ")
teacher_subj = ctk.CTkLabel(Professor_details, text="SUBJECT: ")
teacher_sem = ctk.CTkLabel(Professor_details, text="SEMESTER: ")

teacher_name.grid(row=1, column=0, sticky="w", padx=(10,0))
teacher_subj.grid(row=2, column=0, sticky="w", padx=(10,0))
teacher_sem.grid(row=3, column=0, sticky="w", padx=(10,0))

teacher_name_label = ctk.CTkLabel(Professor_details, text="N/A")
subject_name_label = ctk.CTkLabel(Professor_details, text="N/A")
semester_label = ctk.CTkLabel(Professor_details, text="N/A")

teacher_name_label.grid(row=1, column=1, sticky="w")
subject_name_label.grid(row=2, column=1, sticky="w")
semester_label.grid(row=3, column=1, sticky="w")

# Button Row
buttonFrame = ctk.CTkFrame(left_side)
buttonFrame.grid(row=1, column=0, sticky="w", padx=10)

button_edit = ctk.CTkButton(buttonFrame, text="Edit Student", command=student_edit)
button_add = ctk.CTkButton(buttonFrame, text="Add Student", command=student_add)
button_delete = ctk.CTkButton(buttonFrame, text="Del Student", command=student_delete)

button_edit.grid(row=0, column=0, sticky="w", padx=(10, 2), pady=10)
button_add.grid(row=0, column=1, sticky="w", padx=12, pady=5)
button_delete.grid(row=0, column=3, sticky="w", padx=(2, 10), pady=10)

# Student Details Label
Student_details = ctk.CTkFrame(left_side)
Student_details.grid(row=2, column=0, padx=10, pady=10, sticky="nsew")

details_title2 = ctk.CTkFrame(Student_details)
details_title2.grid(row=0, column=0, sticky="w", padx=10,pady=10)
student_details = ctk.CTkLabel(details_title2, text="STUDENT DETAILS")
student_details.grid(row=0, column=0, sticky="w", padx=10)
Student_name = ctk.CTkLabel(Student_details, text="STUDENT NAME: ")
Student_Status = ctk.CTkLabel(Student_details, text="STUDENT STATUS: ")
Course = ctk.CTkLabel(Student_details, text="COURSE: ")
Year = ctk.CTkLabel(Student_details, text="YEAR: ")
Section = ctk.CTkLabel(Student_details, text="SECTION: ")
FirstQuarter = ctk.CTkLabel(Student_details, text="1ST QUARTER: ")
SecQuarter = ctk.CTkLabel(Student_details, text="2ND QUARTER: ")
Final_grade = ctk.CTkLabel(Student_details, text="FINAL GRADE: ")
Remarks = ctk.CTkLabel(Student_details, text="REMARKS")

Student_name.grid(row=1, column=0, sticky="w", padx=10)
Student_Status.grid(row=2, column=0, sticky="w", padx=10)
Course.grid(row=3, column=0, sticky="w", padx=10)
Year.grid(row=4, column=0, sticky="w", padx=10)
Section.grid(row=5, column=0, sticky="w", padx=10)
FirstQuarter.grid(row=6, column=0, sticky="w", padx=10)
SecQuarter.grid(row=7, column=0, sticky="w", padx=10)
Final_grade.grid(row=8, column=0, sticky="w", padx=10)
Remarks.grid(row=9, column=0, sticky="w", padx=10)

Student_name_label = ctk.CTkLabel(Student_details, text="N/A")
Student_Status_label = ctk.CTkLabel(Student_details, text="N/A")
Course_label = ctk.CTkLabel(Student_details, text="N/A")
Year_label = ctk.CTkLabel(Student_details, text="N/A")
Section_label = ctk.CTkLabel(Student_details, text="N/A")
FirstQuarter_label = ctk.CTkLabel(Student_details, text="N/A")
SecQuarter_label = ctk.CTkLabel(Student_details, text="N/A")
Final_grade_label = ctk.CTkLabel(Student_details, text="N/A")
Remarks_label = ctk.CTkLabel(Student_details, text="N/A")

Student_name_label.grid(row=1, column=1, sticky="w")
Student_Status_label.grid(row=2, column=1, sticky="w")
Course_label.grid(row=3, column=1, sticky="w")
Year_label.grid(row=4, column=1, sticky="w")
Section_label.grid(row=5, column=1, sticky="w")
FirstQuarter_label.grid(row=6, column=1, sticky="w")
SecQuarter_label.grid(row=7, column=1, sticky="w")
Final_grade_label.grid(row=8, column=1, sticky="w")
Remarks_label.grid(row=9, column=1, sticky="w")

def toggle_sem(toggle_id):
    if toggle_id == 1:
        Student_Activity.grid_remove()
        Student_Activity2.grid(row=3, column=0, padx=10, pady=(0,10), sticky="nsew")
    elif toggle_id == 2:
        Student_Activity2.grid_remove()
        Student_Activity.grid(row=3, column=0, padx=10, pady=(0,10), sticky="nsew")
    
# Student Details Activity Label Midterm semester
Student_Activity = ctk.CTkFrame(left_side)
Student_Activity.grid(row=3, column=0, padx=10, pady=(0,10), sticky="nsew")

details_title3 = ctk.CTkFrame(Student_Activity)
details_title3.grid(row=0, column=0, sticky="w", padx=10)
student_act = ctk.CTkLabel(details_title3, text="STUDENT ACTIVITIES")
student_act.grid(row=0, column=0, sticky="w", padx=10)
student_term = ctk.CTkButton(Student_Activity, text="MID-TERM",command=lambda: toggle_sem(1))
student_term.grid(row=0, column=1, sticky="w", padx=10, pady=10)
edit_act = ctk.CTkButton(Student_Activity, text="EDIT ACTIVITIES", command=activity_edit_midterm)
edit_act.grid(row=0, column=2, sticky="nsew", pady=10)
Quiz1 = ctk.CTkLabel(Student_Activity, text="QUIZ #1: ")
Quiz2 = ctk.CTkLabel(Student_Activity, text="QUIZ #2: ")
Quiz3 = ctk.CTkLabel(Student_Activity, text="QUIZ #3: ")
Performance1 = ctk.CTkLabel(Student_Activity, text="PERFORMANCE #1: ")
Performance2 = ctk.CTkLabel(Student_Activity, text="PERFORMANCE #2: ")
Performance3 = ctk.CTkLabel(Student_Activity, text="PERFORMANCE #3: ")
MIDTERM_P = ctk.CTkLabel(Student_Activity, text="MIDTERM PROJECT: ")
MIDTERM_E = ctk.CTkLabel(Student_Activity, text="MIDTERM EXAM: ")
Attendance_Midterm = ctk.CTkLabel(Student_Activity, text="ATTENDANCE MIDTERM: ")

Quiz1.grid(row=1, column=0, sticky="w", padx=10)
Quiz2.grid(row=2, column=0, sticky="w", padx=10)
Quiz3.grid(row=3, column=0, sticky="w", padx=10)
Performance1.grid(row=4, column=0, sticky="w", padx=10)
Performance2.grid(row=5, column=0, sticky="w", padx=10)
Performance3.grid(row=6, column=0, sticky="w", padx=10)
MIDTERM_P.grid(row=7, column=0, sticky="w", padx=10)
MIDTERM_E.grid(row=8, column=0, sticky="w", padx=10)
Attendance_Midterm.grid(row=9, column=0, sticky="w", padx=10)

# Activities Label
q1 = ctk.CTkLabel(Student_Activity, text="N/A")
q2 = ctk.CTkLabel(Student_Activity, text="N/A")
q3 = ctk.CTkLabel(Student_Activity, text="N/A")
p1 = ctk.CTkLabel(Student_Activity, text="N/A")
p2 = ctk.CTkLabel(Student_Activity, text="N/A")
p3 = ctk.CTkLabel(Student_Activity, text="N/A")
mp = ctk.CTkLabel(Student_Activity, text="N/A")
me = ctk.CTkLabel(Student_Activity, text="N/A")
A1 = ctk.CTkLabel(Student_Activity, text="N/A")


q1.grid(row=1, column=1, sticky="e")
q2.grid(row=2, column=1, sticky="e")
q3.grid(row=3, column=1, sticky="e")
p1.grid(row=4, column=1, sticky="e")
p2.grid(row=5, column=1, sticky="e")
p3.grid(row=6, column=1, sticky="e")
mp.grid(row=7, column=1, sticky="e")
me.grid(row=8, column=1, sticky="e")
A1.grid(row=9, column=1, sticky="e")


tq1 = ctk.CTkLabel(Student_Activity, text=" / 10")
tq2 = ctk.CTkLabel(Student_Activity, text=" / 10")
tq3 = ctk.CTkLabel(Student_Activity, text=" / 10")
tp1 = ctk.CTkLabel(Student_Activity, text=" / 50")
tp2 = ctk.CTkLabel(Student_Activity, text=" / 50")
tp3 = ctk.CTkLabel(Student_Activity, text=" / 50")
tmp = ctk.CTkLabel(Student_Activity, text=" / 100")
tme = ctk.CTkLabel(Student_Activity, text=" / 50")
tA1 = ctk.CTkLabel(Student_Activity, text=" / 30")

tq1.grid(row=1, column=2, sticky="w")
tq2.grid(row=2, column=2, sticky="w")
tq3.grid(row=3, column=2, sticky="w")
tp1.grid(row=4, column=2, sticky="w")
tp2.grid(row=5, column=2, sticky="w")
tp3.grid(row=6, column=2, sticky="w")
tmp.grid(row=7, column=2, sticky="w")
tme.grid(row=8, column=2, sticky="w")
tA1.grid(row=9, column=2, sticky="w")
#================================================================================================

Student_Activity2 = ctk.CTkFrame(left_side)
Student_Activity2.grid(row=3, column=0, padx=10, pady=(0,10), sticky="nsew")

details_title4 = ctk.CTkFrame(Student_Activity2)
details_title4.grid(row=0, column=0, sticky="w", padx=10)
student_act = ctk.CTkLabel(details_title4, text="STUDENT ACTIVITIES")
student_act.grid(row=0, column=0, sticky="w", padx=10)
student_term = ctk.CTkButton(Student_Activity2, text="FINAL-TERM",command=lambda: toggle_sem(2))
student_term.grid(row=0, column=1, sticky="w", padx=10, pady=10)
edit_act = ctk.CTkButton(Student_Activity2, text="EDIT ACTIVITIES", command=activity_edit_final)
edit_act.grid(row=0, column=2, sticky="nsew", pady=10)
Quiz1 = ctk.CTkLabel(Student_Activity2, text="QUIZ #1: ")
Quiz2 = ctk.CTkLabel(Student_Activity2, text="QUIZ #2: ")
Quiz3 = ctk.CTkLabel(Student_Activity2, text="QUIZ #3: ")
Performance1 = ctk.CTkLabel(Student_Activity2, text="PERFORMANCE #1: ")
Performance2 = ctk.CTkLabel(Student_Activity2, text="PERFORMANCE #2: ")
Performance3 = ctk.CTkLabel(Student_Activity2, text="PERFORMANCE #3: ")
MIDTERM_P = ctk.CTkLabel(Student_Activity2, text="FINAL PROJECT: ")
MIDTERM_E = ctk.CTkLabel(Student_Activity2, text="FINAL EXAM: ")
Attendance_Midterm = ctk.CTkLabel(Student_Activity2, text="ATTENDANCE FINAL: ")

Quiz1.grid(row=1, column=0, sticky="w", padx=10)
Quiz2.grid(row=2, column=0, sticky="w", padx=10)
Quiz3.grid(row=3, column=0, sticky="w", padx=10)
Performance1.grid(row=4, column=0, sticky="w", padx=10)
Performance2.grid(row=5, column=0, sticky="w", padx=10)
Performance3.grid(row=6, column=0, sticky="w", padx=10)
MIDTERM_P.grid(row=7, column=0, sticky="w", padx=10)
MIDTERM_E.grid(row=8, column=0, sticky="w", padx=10)
Attendance_Midterm.grid(row=9, column=0, sticky="w", padx=10)

# Activities Label
q1f = ctk.CTkLabel(Student_Activity2, text="N/A")
q2f = ctk.CTkLabel(Student_Activity2, text="N/A")
q3f = ctk.CTkLabel(Student_Activity2, text="N/A")
p1f = ctk.CTkLabel(Student_Activity2, text="N/A")
p2f = ctk.CTkLabel(Student_Activity2, text="N/A")
p3f = ctk.CTkLabel(Student_Activity2, text="N/A")
fp = ctk.CTkLabel(Student_Activity2, text="N/A")
fe = ctk.CTkLabel(Student_Activity2, text="N/A")
A1f = ctk.CTkLabel(Student_Activity2, text="N/A")

q1f.grid(row=1, column=1, sticky="e")
q2f.grid(row=2, column=1, sticky="e")
q3f.grid(row=3, column=1, sticky="e")
p1f.grid(row=4, column=1, sticky="e")
p2f.grid(row=5, column=1, sticky="e")
p3f.grid(row=6, column=1, sticky="e")
fp.grid(row=7, column=1, sticky="e")
fe.grid(row=8, column=1, sticky="e")
A1f.grid(row=9, column=1, sticky="e")

tq1f = ctk.CTkLabel(Student_Activity2, text=" / 10")
tq2f = ctk.CTkLabel(Student_Activity2, text=" / 10")
tq3f = ctk.CTkLabel(Student_Activity2, text=" / 10")
tp1f = ctk.CTkLabel(Student_Activity2, text=" / 50")
tp2f = ctk.CTkLabel(Student_Activity2, text=" / 50")
tp3f = ctk.CTkLabel(Student_Activity2, text=" / 50")
tfp = ctk.CTkLabel(Student_Activity2, text=" / 100")
tfe = ctk.CTkLabel(Student_Activity2, text=" / 50")
tA1f = ctk.CTkLabel(Student_Activity2, text=" / 30")

tq1f.grid(row=1, column=2, sticky="w")
tq2f.grid(row=2, column=2, sticky="w")
tq3f.grid(row=3, column=2, sticky="w")
tp1f.grid(row=4, column=2, sticky="w")
tp2f.grid(row=5, column=2, sticky="w")
tp3f.grid(row=6, column=2, sticky="w")
tfp.grid(row=7, column=2, sticky="w")
tfe.grid(row=8, column=2, sticky="w")
tA1f.grid(row=9, column=2, sticky="w")

#========================================================================================

right_side = ctk.CTkFrame(frame)
right_side.grid(row=0, column=1, sticky="nsew")

canvas = tk.Canvas(right_side)
canvas.pack(side="left", fill="both", expand=True)

treeview_frame = ctk.CTkFrame(canvas)
treeview_frame.pack(fill="both", expand=True)

treeview = ttk.Treeview(treeview_frame, show="headings", columns=headings[0:13])
treeview.pack(side="left",fill="both", expand=True)

for col in headings[0:13]:
    treeview.heading(col, text=col, anchor="center")

treeview.column("No.", width=40, minwidth=30, stretch=True)
treeview.column("Student Name", width=120, minwidth=80, stretch=True)
treeview.column("Quiz1", width=50, minwidth=50, stretch=True)
treeview.column("Perf1", width=50, minwidth=50, stretch=True)
treeview.column("Midterm Project", width=100, minwidth=80, stretch=True)
treeview.column("Midterm Exam", width=100, minwidth=80, stretch=True)
treeview.column("Quiz2", width=50, minwidth=50, stretch=True)
treeview.column("Perf2", width=50, minwidth=50, stretch=True)
treeview.column("Finalterm Project", width=100, minwidth=80, stretch=True)
treeview.column("Finalterm Exam", width=100, minwidth=80, stretch=True)
treeview.column("Initial Grade", width=80, minwidth=80, stretch=True)
treeview.column("Final Grade", width=70, minwidth=50, stretch=True)
treeview.column("Total Grade", width=70, minwidth=50, stretch=True)

treeview_scrollbar = ttk.Scrollbar(treeview_frame, orient="vertical", command=treeview.yview)
treeview_scrollbar.pack(side="right", fill="y")

treeview.configure(yscrollcommand=treeview_scrollbar.set)
treeview.bind("<<TreeviewSelect>>", lambda event: on_select(event, treeview))

display_excel_data()

root.mainloop()